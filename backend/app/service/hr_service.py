import csv
import io
from app.extension import db
from app.model.leave_request import LeaveRequest
from app.model.leave_balance import LeaveBalance
from app.model.user import User
from app.model.leave_type import LeaveType


def _ascii(text: str) -> str:
    """Replace Polish diacritics with ASCII equivalents for PDF output."""
    return text.translate(str.maketrans(
        'ąćęłńóśźżĄĆĘŁŃÓŚŹŻ',
        'acelnoszzACELNOSZZ'
    ))


def _query_all_leaves():
    return db.session.query(LeaveRequest, User, LeaveType).join(
        User, LeaveRequest.user_id == User.id
    ).join(
        LeaveType, LeaveRequest.leave_type_id == LeaveType.id
    ).all()


def get_all_leaves() -> list[LeaveRequest]:
    return db.session.query(LeaveRequest).all()


def get_all_leaves_detailed() -> list[dict]:
    results = _query_all_leaves()
    return [
        {
            "id": req.id,
            "user_name": f"{user.first_name} {user.last_name}",
            "email": user.email,
            "leave_type": leave_type.name,
            "start_date": req.start_date.isoformat(),
            "end_date": req.end_date.isoformat(),
            "status": req.status.value,
            "reason": req.request_reason,
        }
        for req, user, leave_type in results
    ]


def generate_leaves_csv_report() -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Request ID', 'Employee', 'Email', 'Leave Type', 'Start Date', 'End Date', 'Status'])
    for request, user, leave_type in _query_all_leaves():
        writer.writerow([
            request.id,
            f"{user.first_name} {user.last_name}",
            user.email,
            leave_type.name,
            request.start_date.isoformat(),
            request.end_date.isoformat(),
            request.status.value
        ])
    return output.getvalue()


def generate_leaves_pdf_report() -> bytes:
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Raport Urlopowy", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 9)
    cols = [("ID", 10), ("Pracownik", 42), ("Email", 52), ("Typ", 28), ("Od", 22), ("Do", 22), ("Status", 25)]
    for header, width in cols:
        pdf.cell(width, 7, header, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", size=8)
    for req, user, leave_type in _query_all_leaves():
        row = [
            (str(req.id), 10),
            (_ascii(f"{user.first_name} {user.last_name}")[:22], 42),
            (_ascii(user.email)[:30], 52),
            (_ascii(leave_type.name)[:18], 28),
            (req.start_date.isoformat(), 22),
            (req.end_date.isoformat(), 22),
            (req.status.value, 25),
        ]
        for value, width in row:
            pdf.cell(width, 6, value, border=1)
        pdf.ln()

    return bytes(pdf.output())


def get_all_balances() -> list[dict]:
    results = db.session.query(LeaveBalance, User, LeaveType).join(
        User, LeaveBalance.user_id == User.id
    ).join(
        LeaveType, LeaveBalance.leave_type_id == LeaveType.id
    ).order_by(User.last_name, LeaveType.name).all()

    return [
        {
            "id": b.id,
            "user_id": b.user_id,
            "user_name": f"{u.first_name} {u.last_name}",
            "leave_type_id": b.leave_type_id,
            "leave_type_name": lt.name,
            "year": b.year,
            "total_days": b.total_days,
            "used_days": b.used_days,
            "remaining_days": b.total_days - b.used_days,
        }
        for b, u, lt in results
    ]


def create_balance_for_user(user_id: int, leave_type_id: int, year: int, total_days: int) -> LeaveBalance:
    existing = db.session.query(LeaveBalance).filter_by(
        user_id=user_id, leave_type_id=leave_type_id, year=year
    ).first()

    if existing:
        raise ValueError("Balance already exists for this user, leave type, and year.")

    balance = LeaveBalance(
        user_id=user_id,
        leave_type_id=leave_type_id,
        year=year,
        total_days=total_days,
        used_days=0,
    )
    db.session.add(balance)
    db.session.commit()
    return balance


def update_balance_total(balance_id: int, total_days: int) -> LeaveBalance:
    balance = db.session.get(LeaveBalance, balance_id)
    if not balance:
        raise ValueError("Balance not found.")

    if total_days < balance.used_days:
        raise ValueError(
            f"Cannot set total days below already used days ({balance.used_days})."
        )

    balance.total_days = total_days
    db.session.commit()
    return balance


def generate_leaves_xls_report() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raport Urlopowy"
    ws.append(["ID", "Pracownik", "Email", "Typ urlopu", "Data od", "Data do", "Status"])

    for req, user, leave_type in _query_all_leaves():
        ws.append([
            req.id,
            f"{user.first_name} {user.last_name}",
            user.email,
            leave_type.name,
            req.start_date.isoformat(),
            req.end_date.isoformat(),
            req.status.value,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
